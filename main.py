# GOAL
# a time tracker app
# user to select the task from a task list drop down (data source for task list -> excel)
# has buttons -> start, pause, end, reset
# stores the start time, end time in Excel on click of end button
# resets timer on reset button click
import time
import customtkinter as ctk
import os
import pandas as pd
from PIL.ImageFile import ImageFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import datetime as dt
from enum import Enum
# for handling icons
# ImageTk for toolbar icon in about and manage task status windows
from PIL import Image, ImageDraw, ImageTk
# for handling file paths
import sys
# to create a separate thread for sys tray icon
import threading
# to create a sys tray icon and to create menu items for sys tray icon right click
import pystray
# to get dpi scaling
from ctypes import windll


class TimerStatus(Enum):
    # to track the timer status for the buttons to work correctly
    RUNNING = 1
    PAUSED = 2
    STOPPED = 3

class TaskTimer:
    def __init__(self) -> None:
        # set the window theme to 'dark' mode
        ctk.set_appearance_mode("dark")
        # initialize the main window
        self.app = ctk.CTk()
        self.app_title = "Time Keeper"
        self.app.title(self.app_title)
        # disable window resizing
        self.app.resizable(False, False)
        # to disable the toolbar and make it as a widget,
        # makes a window borderless and removes from taskbar
        self.app.overrideredirect(True)

        # -----------assets-----------
        # Excel file to store the task list, time
        self.excel_file = "Time_Keeper.xlsx"
        # icon for the system tray
        self.app_icon = "app_icon.ico"
        # Sheet in the Excel file to store the task list
        self.excel_tasks_sheet = "Tasks"
        # name of the column storing tasks inside the Tasks sheet
        self.tasks_col_name = "Task"
        # Sheet in the Excel file to store the time for each task
        self.excel_time_sheet = "Time"
        # excel icon for excel_btn
        self.excel_btn_icon = "excel_btn_icon.png"
        self.task_active_status_symbol = "Active"


        # To track the work duration for the current date
        # We use .replace as only this works in _get_days_work_minutes() method
        self.current_date = dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        self.days_work_minutes = self._get_days_work_minutes()

        # get the task list from the Excel if it exists, to populate task_list_menu combobox dropdown
        self.all_tasks_dict_list = None
        # this is only active tasks list
        self.task_list = self._get_task_list()
        # task selected from the task_list_menu combobox
        self.current_task = ""

        # track the timer status - running, paused, stopped
        self.is_timer_running: TimerStatus = TimerStatus.STOPPED
        # to display timer text inside the timer_display Entry
        self.timer_text = ctk.StringVar()
        # to track the number of seconds elapsed and to use to set the text for timer_display Entry via timer_text
        # self.seconds_elapsed_ui = 0
        self.task_start_time = None # to store the start time of the task
        self.task_end_time = None # to store the end time of the task
        self.segment_start_time = None # start of a segment, if paused
        self.segment_end_time = None # end of a segment, if paused
        self.work_seconds = 0
        # ----to track multi day task tracking variables----
        self.new_day_pause_start = None
        self.new_day_pause_seconds = 0
        self.work_seconds_logged = 0
        self.multiday_start_date = None # to group multiday tasks

        # to manage placeholder text in the notes_entry field
        # when is_placeholder_active is True, show PH text in the notes_entry filed
        self.is_placeholder_active = True

        # declared it here as these are used by multiple methods
        self.days_work_label = None
        self.task_list_menu = None
        self.status_label = None
        self.timer_display = None
        self.start_btn = None
        # Textbox for user to type in task notes
        self.notes_textbox = None
        # To be disabled when timer is running or paused
        self.manage_tasks_btn = None

        # to store the after() ID and to handle .after() calls overlaps i.e., to be used in .after_cancel()
        self.status_update_queue = None
        self.timer_running_queue = None

        # to handle drag and reposition of the app window
        self.start_mouse_x_root = None
        self.start_mouse_y_root = None
        self.start_window_x_root = None
        self.start_window_y_root = None

        # to enable running and controlling the app from the system tray
        self.systray_icon = None
        # create a separate thread for sys tray icon run so that mainloop() does not block this
        self.systray_thread = threading.Thread(target=self._initialize_systray_icon, daemon=True)
        # start the sys tray thread
        self.systray_thread.start()

        # build the ui (widgets) of the app
        self._build_ui()

        # position app window in the bottom right corner of the screen
        self.position_window()
        # ensure the app is brought to the top on start
        self.app.attributes("-topmost", True)
        # self.app.attributes("-topmost", False)
        # retain the top position for 500 ms and release after that
        self.app.after(500, lambda: self.app.attributes('-topmost', False))

        # to track the system sleep/freeze/hang phases etc.
        self.last_ui_update_mono = time.monotonic()
        self.last_ui_update_time = dt.datetime.now()
        # perpetual loop to log last UI update time to detect system sleep/freeze/hang etc.
        self._schedule_update_timer()
        # to track if the end of the task is by user of by system so that the task_end_time and segment_end_time are set accordingly
        self.auto_end = False

        # start the loop to check if day has changed and update the day's duration display
        self.check_day_change_queue = self.app.after(60000, self._check_for_day_change_periodically)

        self.app.mainloop()


    def _get_days_work_minutes(self) -> int:
        """
        Get the total of days work minutes from the Excel when the app is opened
        If the Excel does not exist, reruns 0
        :return:
        """
        # 1. check if the Excel file exists
        if os.path.exists(self.excel_file) and os.path.getsize(self.excel_file) > 0:
            # check if the Time sheet exists and catch errors on read
            try:
                # 2. read the Time sheet in the Excel file
                time_df = pd.read_excel(self.excel_file, sheet_name=self.excel_time_sheet)

                # check if DF is not empty (e.g., only headers)
                if not time_df.empty:
                    # get the line items of current date
                    days_time_df = time_df[ time_df["Date"] == self.current_date ]
                    days_work_minutes_list = days_time_df["Work_Minutes"].tolist()
                    days_work_minutes = sum(days_work_minutes_list)
                    return days_work_minutes
            except pd.errors.ParserError as e:
                print(f"Error reading the file to get the Time list: {e}")
            except Exception as e:
                # catch any other unexpected errors
                print(f"An unexpected error occurred while getting the Time list: {e}")

        return 0


    def _check_for_day_change_periodically(self):
        """
        Checks if a new day has started every 1 minute, and if a new day is detected,
        resets the days work minutes to 0 (for UI display) and current date to new day's date
        triggers the previous day's data logging if the timer is not stopped i.e., paused or running
        """
        # get the calendar date and compare it with the date currently we are displaying the day's duration for
        if self.current_date != dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            # a new day has started

            if self.is_timer_running != TimerStatus.STOPPED:
                # if the timer is running or is paused,
                # log the previous day's data to the Excel
                self._check_day_split_and_log()

            # update the UI to display the new day's work minutes, which would mostly be 0
            # _check_day_split_and_log() method also triggers _update_days_work_minutes_display() method
            # but only for current/new day's log and not for the previous day's log
            # so we call _update_days_work_minutes_display() here irrespective of timer running status
            self._update_days_work_minutes_display()



            print(f"{self.current_date=}, {self.days_work_minutes=}")

        # to ensure the check runs perpetually
        self.check_day_change_queue = self.app.after(60000, self._check_for_day_change_periodically)


    def _get_task_list(self):
        """
        Read tasks from the Excel file if it exists and is not empty
        Only includes tasks where Status='Active'
        :returns list: A list of tasks always starting with "<Add new task...>".
        Returns just ["<Add new task...>"] if the Excel file doesn't exist, is empty, or has invalid data.
        """

        tasks_list = []

        # 1. check if the file exists and is not empty
        if os.path.exists(self.excel_file) and os.path.getsize(self.excel_file) > 0:
            # to catch errors in reading the file
            try:
                # 2. read the Excel sheet into a DF
                tasks_df = pd.read_excel(self.excel_file, sheet_name=self.excel_tasks_sheet)
                # print(f"tasks df:\n{tasks_df}")
                # to check if a task exists on new task addition
                self.all_tasks_dict_list = tasks_df.to_dict(orient="records")

                # 3. check if DF is not empty (e.g., only headers)
                if not tasks_df.empty:
                    # 4. drop blanks in 'Tasks' column and filter out tasks with status != 'active'
                    active_tasks = tasks_df[
                        tasks_df[self.tasks_col_name].notna() &
                        (tasks_df["Status"].str.lower() == self.task_active_status_symbol.lower())
                    ]

                    # 5. convert the tasks to a list
                    tasks_list = active_tasks[self.tasks_col_name].to_list()
                    tasks_list.sort()

                else:
                    print(f"{self.excel_file} exists but contains no data.")
            except pd.errors.ParserError as e:
                print(f"Error reading the file to get the task list: {e}")
            except Exception as e:
                # catch any other unexpected errors
                print(f"An unexpected error occurred while getting the task list: {e}")
        else:
            print(f"{self.excel_file} doesn't exist or is empty.")

        # print(f"{tasks_list=}")
        return ["<Add new task...>"] + tasks_list


    def _list_menu_callback(self, choice):
        # instead of removing the existing task manually, if any,
        # user can select the "<Add new task...>" item to clear the field and set the focus to type the new task
        # space before < to ensure this stays at the top after sorting the list
        if choice == "<Add new task...>":
            self.task_list_menu.set("")
            # set the current_task to "" so that the previously set task is cleared, else we can still run the timer without typing or selecting a new task
            # Flow - select a task from the dropdown (task_list_menu combobox) -> current_task = selected_task due to else block
            # select "<Add new task...>" from the dropdown, task_list_menu will be set to "" and cursor will be blinking, but the current_task still has a value from previous selection
            # so, even without typing in a new task or selecting one from the dropdown, use can start the timer with start_btn as the check 'if self.current_task' in run_timer evaluates true
            # hence we have to clear the current_task whenever '"<Add new task...>"' is selected
            self.current_task = ""
            self.task_list_menu.focus()
        else:
            # set the current task value
            self.current_task = choice
            # to remove the blinking cursor in the combobox after item selection
            self.app.focus()
        # print("Selected Task:", choice)


    def _update_status_label(self, status: str, code: int):
        """
        Display status of a task for 3-seconds
        :param status: str
        :param code: int | 0 for success 1 for failure/error/warning
        """
        # check if a status update task is running and cancel it before starting a new task
        if self.status_update_queue is not None:
            self.app.after_cancel(self.status_update_queue)

        # show the task addition status
        if code == 1:
            # if there is an error/failure
            status = status + " :(" if status else status
            self.status_label.configure(text=status, text_color="#b54747")
        else:
            # if the task is successful
            status = status + " :)" if status else status
            self.status_label.configure(text=status, text_color="#009933")

        # schedule a new status update task to hide the status and store the ID
        # schedule task only if the status is not empty, else it will lead to infinite loop
        # Call 1 ("Added") → Call 2 ("") → Call 3 ("") → Call 4 ("") → ...
        if status:
            self.status_update_queue = self.app.after(3000,
                                                      lambda: self._update_status_label("", 1)
                                                      )


    def _append_data_to_excel(self, sheet_name, **kwargs) -> bool:
        """
        Appends new row of data to the specified sheet in the Excel file
        Creates the Excel file/new sheet if they don't exist
        Data is passed as keyword arguments and keywords become headers
        Example usage:
            _append_data_to_excel("Tasks", Task="Study Python", Status="Active", Added_On="2025-04-05 10:00")
            _append_data_to_excel("Time", Task="Study Python", Timestamp="2025-04-05 10:00", Notes="Great progress!")
        :param sheet_name: (str): Name of the sheet to append to
        :param kwargs: Each key becomes a column header, value becomes cell data
        :returns bool: True if successful, False otherwise
        """
        # using openpyxl for appending data as calculating last row (with openpyxl) is required for pandas
        try:
            # 1. check if the file exists or to be created
            if os.path.exists(self.excel_file):
                try:
                    # Attempt to load the workbook. This will fail for zero-byte or corrupted files
                    wb = load_workbook(self.excel_file)
                except (InvalidFileException, Exception) as e:
                    print(f"Error with existing file: {e}. Creating new file.")
                    wb = Workbook()
                    # remove default sheets created e.g., 'Sheet1'
                    if len(wb.sheetnames) > 0:
                        for s in wb.sheetnames:
                            wb.remove(wb[s])
            else:
                # file does not exist, so creating a new file
                wb = Workbook()
                # remove default sheets created e.g., 'Sheet1'
                if len(wb.sheetnames) > 0:
                    for s in wb.sheetnames:
                        wb.remove(wb[s])

            # 2. check if the sheet exists or to be created
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
                # set the headings for the sheet
                sheet.append(list(kwargs.keys()))
            else:
                # sheet exists in the Excel file
                sheet = wb[sheet_name]

            # 3. append the new data
            sheet.append( list(kwargs.values()) )

            # 4. save and close the Excel file
            wb.save(self.excel_file)
            wb.close()
            return True
        except Exception as e:
            # This outer catch is for errors during sheet creation, appending, or saving
            print(f"Error on appending data to excel: {e}")
            return False


    def _show_placeholder(self):
        """
        sets the placeholder text in the notes_entry field
        """
        self.notes_textbox.insert("1.0", "Add notes")
        self.notes_textbox.configure(text_color="#7a848d")
        # we set the status here instead of in _notes_focus_out() method so that when we call this method in _reset_timer(), even the status is also set
        self.is_placeholder_active = True


    def _notes_focus_in(self, event):
        """
        when notes_textbox is focused and if is_placeholder_active = True
        clear the placeholder text and change font color #7a848d -> #f2f2f2
        """
        if self.is_placeholder_active:
            self.notes_textbox.delete("1.0", "end")
            # abb9c6 afb4ba
            self.notes_textbox.configure(text_color="#d2d9e0")
            self.is_placeholder_active = False


    def _notes_focus_out(self, event):
        """
        when notes_textbox loses focus, checks if there is a user entered text in it
        if there is no user entered text, then show the placeholder
        """
        if not self.notes_textbox.get("1.0", "end-1c").strip():
            # user has not entered any text or entered just spaces
            self._show_placeholder()
            # self.is_placeholder_active = True


    def _add_task_on_enter(self, event):
        """
        To add a new task typed into the task_list_menu combobox to the task_list and Excel, on press of 'Enter' key
        """
        # get the text currently in the combobox entry
        # .strip() removes leading/trailing whitespace
        new_task = self.task_list_menu.get().strip()

        # if the new_task is not empty and does not exist in the task_list, add it to the task_list
        if new_task:
            # to preserve formats like 'ITR' 'GPS'
            new_task = new_task[0].upper() + new_task[1:]
            # check if there are any existing tasks and then check if the new task exists
            if self.all_tasks_dict_list:
                # For robust check on if the task exists,
                does_task_exist = any(task_item[self.tasks_col_name].lower() == new_task.lower()
                                  for task_item in self.all_tasks_dict_list)
            else:
                does_task_exist = False

            if not does_task_exist:
                # add/append the task to excel and if that is successful, proceed further
                now = dt.datetime.now()
                write_status = self._append_data_to_excel(self.excel_tasks_sheet, Task=new_task,
                                                          Status=self.task_active_status_symbol,
                                                          Added_On=f"{now:%d-%b-%Y T%I:%M %p}")

                # perform further steps
                if write_status:
                    self.task_list = self._get_task_list()
                    # to ensure "<Add new task...>" is at the top of the list
                    self.task_list[1:] = sorted(self.task_list[1:])
                    # update the combobox with the new task_list
                    self.task_list_menu.configure(values=self.task_list)
                    # set the value to new_task with spaces stripped and capitalized
                    self.task_list_menu.set(new_task)
                    # update the current task selection which will be used as validation for starting timer on click of start_btn in run_timer method
                    self.current_task = new_task
                    # show the task addition status
                    self._update_status_label("Added", 0)
                    # to remove focus (cursor) from the task_list_menu combobox
                    self.app.focus()
                else:
                    self._update_status_label("Error", 1)
            else:
                self._update_status_label("Exists", 1)
        else:
            self._update_status_label("Empty", 1)


    def _update_timer_display(self):
        """
        updates the timer display every second if the timer is RUNNING
        detects system sleep/freeze and if detected, will end the current timer and log the data to Excel by calling the end_timer() method
        """
        # UI should update every 1000ms due to .after() calls
        # 0.5 buffer to address scheduling delays
        update_interval = 1.5

        if self.is_timer_running == TimerStatus.RUNNING:
            current_mono = time.monotonic()

            time_since_last_ui_update = current_mono - self.last_ui_update_mono

            if time_since_last_ui_update < update_interval:
                # the system did not sleep, or there was no UI freeze
                # use .seconds instead of .total_seconds() as the later keeps accumulating the fractional seconds that may lead to a jump between multiple pause and resume cycles (this happens if we include decimal points also that we get with .total_seconds(). but we exclude the decimal part with int() )
                # .seconds only gives max 86400 i.e., seconds for the day, so use total_seconds() and int()
                # Why we do not accumulate work_seconds here and use _seconds_aacumulator() for that?
                #
                seconds_elapsed_ui = self.work_seconds + (dt.datetime.now() -
                                                          self.segment_start_time).total_seconds()
                seconds_elapsed_ui = int(seconds_elapsed_ui)
                # print(f"_update_timer_display() -> {self.work_seconds=} {seconds_elapsed_ui=}")
                hours_elapsed, remainder = divmod(seconds_elapsed_ui, 3600)
                # the remainder we get here is the seconds remaining
                minutes_elapsed, remainder = divmod(remainder, 60)
                # show the timer in the timer_display Entry via timer_text instance variable
                self.timer_text.set(f"{hours_elapsed:02}:{minutes_elapsed:02}:{remainder:02}")
            else:
                # the system is awake from sleep or recovered from a freeze/hang
                # in this case, the task and segment were running till the last_ui_update_time
                # we accumulate work from the segment that was running before sleep
                # till the last known active moment before suspension i.e., last_ui_update
                self.segment_end_time = self.last_ui_update_time
                # we set the task end time to last_ui_update_time so that sleep time is excluded as we use task_endTime to calculate pause duration inside _calculate_duration() method
                self.task_end_time = self.last_ui_update_time
                self.auto_end = True
                self._seconds_accumulator()
                self._end_timer()

        # capture the last ui update time to check for system sleep/freeze by calculating the diff between this and next ui update time
        self.last_ui_update_time = dt.datetime.now()
        self.last_ui_update_mono = time.monotonic()
        # call the schedule timer method irrespective of the timer status to ensure perpetual loop for sleep/freeze detection
        self._schedule_update_timer()


    def _schedule_update_timer(self):
        # perpetual loop to detect system sleep/freeze
       self.timer_running_queue=  self.app.after(1000, self._update_timer_display)


    def _humanize_time(self, minutes):
        """
        Formats the total minutes to H:MM string format to be logged to Excel
        Handles durations longer than 24 hours by accumulating the hours
        :return: str: 2h 12m or 0
        """
        # this is incorrect as there may be pause time in between
        # difference = self.task_end_time - self.task_start_time
        # diff_seconds = difference.total_seconds()
        # seconds_elapsed represents only the time in seconds timer ran and not paused
        if minutes:
            hours, minutes = divmod(minutes, 60)
            return f"{hours:.0f}h {minutes:02.0f}m" if hours else f"{minutes:.0f}m"
        else:
            return 0


    def _seconds_accumulator(self):
        """
        calculate the segment duration (segment_end_time - segment_start_time), when paused
        and add the duration to work seconds
        only completed second is considered i.e., numbers after decimal. that we get from .total_seconds() is ignored
        """
        self.work_seconds += int((self.segment_end_time - self.segment_start_time).total_seconds())


    def _new_day_pause_seconds_accumulator(self, current_timestamp: dt.datetime) -> None:
        """
        Accumulates the pause seconds of a new day for multi-day tasks
        :return: None
        """
        # if pause start is captured (not None), pause start and end or on a new day, else pause start is on the previous day (i.e., before midnight) and end (the pause end = resume) is on new day
        if self.new_day_pause_start:
            # pause start was on a new day
            self.new_day_pause_seconds += int( (current_timestamp - self.new_day_pause_start).total_seconds())
            self.new_day_pause_start = None
        else:
            # pause start was on the previous day
            midnight_timestamp = dt.datetime.combine(current_timestamp, dt.time())
            # new day pause duration will be midnight -> resume time
            self.new_day_pause_seconds += int( (current_timestamp - midnight_timestamp).total_seconds() )


    def _run_timer(self):
        """
        Handles starting, pausing, resuming timer
        """
        # if self.start_btn.cget("text") == "▶":
        # check if a task is selected before starting the timer
        if self.current_task:
            # To prevent active task status change,
            # We do this instead of setting the state as we can't control the disabled state text color
            self.manage_tasks_btn.configure(command=lambda: ..., text_color="#353535")
            # run timer if the timer is not running i.e., timer is paused or stopped
            if self.is_timer_running != TimerStatus.RUNNING:
                # we set the task_start_time before changing the is_timer_running status to catch the scenario where a timer that was paused is being resumed
                # when paused, is_timer_running = PAUSED, so != RUNNING
                # we then set is_timer_running = RUNNING and after that we again set task_start_time as this inner if evaluates true,
                # i.e., RUNNING != PAUSED which means a new task_start_time is created for an already running task
                # Set task_start_time only if the timer is beginning a new session (from STOPPED);
                # this preserves the original start time when resuming from a PAUSED state.
                # only set if completely new task
                if self.is_timer_running == TimerStatus.STOPPED:
                    # set the task_start_time at the start of the task
                    self.task_start_time = dt.datetime.now()

                # set the start of the segment
                self.segment_start_time = dt.datetime.now()

                # accumulate the pause seconds if new day on resuming a pause
                # check if the resume (i.e., status was PAUSED before resume) is on a new day
                if self.is_timer_running == TimerStatus.PAUSED and self.task_start_time.date() != self.segment_start_time.date():
                    self._new_day_pause_seconds_accumulator(current_timestamp= self.segment_start_time)

                self.is_timer_running = TimerStatus.RUNNING
                self.start_btn.configure(text="⏸")
                # to not select a new task while the timer is running
                self.task_list_menu.configure(state="disabled")
                self._update_status_label("Start", 0)
                # print("Timer running")
            else:
                # timer is paused
                # capture the end of segment when paused
                self.segment_end_time = dt.datetime.now()
                if self.task_start_time.date() != self.segment_end_time.date():
                    # paused on a new day, capture pause start for calculating pause duration
                    self.new_day_pause_start = dt.datetime.now()
                # accumulate the work seconds when paused
                self._seconds_accumulator()
                self.is_timer_running = TimerStatus.PAUSED
                self.start_btn.configure(text="▶")
                # as we are updating self.is_timer_running = TimerStatus.STOPPED, _update_timer() will stop as it runs only when the timer is running -> if self.is_timer_running == TimerStatus.RUNNING: ...
                # self._update_timer()
                self._update_status_label("Pause", 0)
                # print("Timer paused")
        else:
            # if no task is selected before starting the timer (hitting start_btn)
            self._update_status_label("Select", 1)
        self.app.focus()


    def _calculate_duration(self, task_start_time, task_end_time, work_seconds):
        """
        calculates work_minutes, pause_minutes, and total_task_minutes at the end of the task to be saved to Excel
        ensures work_minutes + pause_minutes == total_task_minutes, to align with start and end time shown in hh:mm in Excel
        """

        # we have to tie in task start and end time, trimmed task start and end time, work_minutes, pause_minutes, total_minutes together to ensure total_m = work_m + pause_m

        # trim the seconds for Excel visible times
        task_start_trimmed = task_start_time.replace(second=0, microsecond=0)
        task_end_trimmed = task_end_time.replace(second=0, microsecond=0)

        total_task_seconds_trimmed = int( (task_end_trimmed - task_start_trimmed).total_seconds() )
        total_task_minutes = total_task_seconds_trimmed // 60

        # calculate total wall-clock seconds for the task
        # cast to int for consistency with work_seconds
        total_actual_seconds = int( (task_end_time - task_start_time).total_seconds() )

        # Since we accumulate the work_seconds at hh:mm:ss level, i.e., at seconds precision,
        # we also get the pause duration at seconds precision
        # max to cover cases where work_seconds may be more than total_actual_seconds
        pause_seconds = max(0, total_actual_seconds - work_seconds)
        pause_minutes = pause_seconds // 60

        # derive work_minutes to ensure the work_m + pause_m = total_m
        work_minutes = total_task_minutes - pause_minutes
        work_minutes = max(0, work_minutes)  # Ensure non-negative

        return work_minutes, pause_minutes, total_task_minutes


    # def _dev_log_data(self):
    #     work_minutes, pause_minutes, total_task_minutes = self._calculate_duration()
    #
    #     write_status = self._append_data_to_excel(
    #         sheet_name="Log",
    #         Date=f"{self.task_start_time:%d-%b-%Y}",
    #         Task=self.current_task,
    #         Start=f"{self.task_start_time:%I:%M:%S}",
    #         End=f"{self.task_end_time:%I:%M:%S}",
    #         Total_Seconds = int((self.task_end_time - self.task_start_time).total_seconds()),
    #         Work_Seconds = self.work_seconds,
    #         Work_Minutes = work_minutes,
    #         Pause_Minutes = pause_minutes,
    #         Total_Minutes = total_task_minutes
    #     )


    def _update_days_work_minutes_display(self, current_task_work_minutes=0) -> None:
        """
        Update the UI to show the latest 'days work duration' after the end of the end of a task
        :return: None
        """
        # this is run on task end and triggered from _end_timer() -> log data to excel method
        # or on detection of a date change from _check_for_day_change_periodically()
        # check if the current date is equal to task end date i.e., task start and end are on the same date
        # if yes, add to days work minutes
        # if not, reset the days work minutes and change the current date
        if self.current_date == dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0):
            self.days_work_minutes += current_task_work_minutes
        else:
            self.current_date = dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            # this call to _get_days_work_minutes() may not be required, but we do it be safe
            self.days_work_minutes = self._get_days_work_minutes()

        days_work_minutes_formated = self._humanize_time(self.days_work_minutes)
        self.days_work_label.configure(text=f"Day: {days_work_minutes_formated}")


    def _log_data_to_excel(self, task_start_time, task_end_time, work_minutes, pause_minutes, total_task_minutes) -> bool:
        """
        Logs data to Excel on end of the timer
        Triggered by _end_timer() method
        :return: log_status (True or False)
        """

        # get the work duration in hh:mm format
        work_duration = self._humanize_time(work_minutes)
        pause_duration = self._humanize_time(pause_minutes)

        # to avoid capturing placeholder text as notes
        if self.is_placeholder_active:
            task_notes = ""
        else:
            task_notes = self.notes_textbox.get("1.0", "end-1c")

        # write data to the Excel Date, Task, Duration, Notes, Start Time, End Time, Seconds
        # print(f"{task_start_time.date()=},{self.current_task=}, {work_duration=},{task_notes=},{pause_duration=}, {task_start_time:%I:%M %p}, {task_end_time:%I:%M %p}, {work_minutes=}, {pause_minutes=}, {total_task_minutes=}")
        log_status = self._append_data_to_excel(self.excel_time_sheet,
                                                Date=task_start_time.date(),
                                                Task=self.current_task,
                                                Work_Duration=work_duration,
                                                Notes=task_notes,
                                                Pause_Duration=pause_duration,
                                                Start_Time=f"{task_start_time:%I:%M %p}",
                                                End_Time=f"{task_end_time:%I:%M %p}",
                                                Work_Minutes=work_minutes,
                                                Pause_Minutes=pause_minutes,
                                                Total_Minutes=total_task_minutes,
                                                Multi_day_Start = f"{self.multiday_start_date}",
                                                )

        return log_status


    def _check_day_split_and_log(self) -> bool:
        """
        If the task spans for more than one day, logs entry for each day splitting the duration till midnight
        Else, logs the task for the day
        This method is called on detection of new day start inside _check_for_day_change_periodically()
        Or on task end inside _end_timer()
        :return: Log Status as bool
        """

        # check if task start date and now are on different dates
        current_timestamp = dt.datetime.now() # time when a new day is detected or when the task stopped/ended by the user
        if self.task_start_time.date() != current_timestamp.date():
            # this block is triggered from both _check_for_day_change_periodically() and _end_timer()

            # a new day has started
            cumulative_work_seconds = self.work_seconds

            if self.is_timer_running == TimerStatus.RUNNING:
                # the _seconds_accumulator() would not have been called for the current segment
                cumulative_work_seconds += int( (current_timestamp - self.segment_start_time).total_seconds() )

            # CASE - pause may have started on previous day/new day and not resumed
            # new day change gets triggered inside _check_for_day_change_periodically that calls _check_day_split_and_log, but pause duration accumulation does not happen in run_timer
            # so, we have to count the pause duration from pause start to current_timestamp
            if self.is_timer_running == TimerStatus.PAUSED:
                self._new_day_pause_seconds_accumulator(current_timestamp=current_timestamp)

            # midnight -> now i.e., when a new day was detected/when the task ended on new day
            midnight_timestamp = dt.datetime.combine(current_timestamp, dt.time())
            new_day_total_seconds = int( (current_timestamp - midnight_timestamp).total_seconds() )
            new_day_work_seconds = new_day_total_seconds - self.new_day_pause_seconds

            prev_day_work_seconds = cumulative_work_seconds - new_day_work_seconds

            # to track the multi-day tasks
            self.multiday_start_date = f"{self.task_start_time.date()}"

            # calculate the durations for the previous day i.e., task_start -> midnight
            work_minutes, pause_minutes, total_task_minutes = self._calculate_duration(
                                                                task_start_time=self.task_start_time,
                                                                task_end_time=midnight_timestamp,
                                                                work_seconds=prev_day_work_seconds)

            # log the previous day to the Excel
            prev_day_log_status = self._log_data_to_excel(task_start_time=self.task_start_time,
                                                          task_end_time=midnight_timestamp,
                                                          work_minutes=work_minutes,
                                                          pause_minutes=pause_minutes,
                                                          total_task_minutes=total_task_minutes)

            if prev_day_log_status:
                # if the previous day is successfully logged to Excel, reset the variables
                # we keep adding to work_seconds_logged to handle tasks spanning 2+ days
                self.work_seconds_logged += prev_day_work_seconds
                self.task_start_time = midnight_timestamp
                # Multi - day tracking variables reset
                # we reset these here instead of in _reset_timer() as _check_day_split_and_log() is also called from _check_for_day_change_periodically() where _reset_timer() is no where triggered
                self.new_day_pause_start = None
                self.new_day_pause_seconds = 0
            else:
                return False

        # CASE - task started on the previous day and ended on new day
        # on end, _end_timer() is triggered, which in turn calls this method to log both previous day and new day data
        # if the previous day data log is unsuccessful we should not proceed for new day log
        # CASE - task start and end are on the same day
        # in this case, there is no previous day log status, and we only have to log the current day
        # to handle these situations, we have a return False statement above

        day_log_status = ""
        if self.is_timer_running == TimerStatus.STOPPED:
            # this block is triggered on call from _end_timer() where we have to log for
            # only current day and the new day
            # previous day is handled by the above if block

            # get the current day's work seconds (matters if the task is a multi-day task)
            current_day_work_seconds = self.work_seconds - self.work_seconds_logged

            # calculate the durations for the current day or new day (if multi-day task)
            work_minutes, pause_minutes, total_task_minutes = self._calculate_duration(
                                                                    task_start_time=self.task_start_time,
                                                                    task_end_time=self.task_end_time,
                                                                    work_seconds=current_day_work_seconds)

            # log the previous day to the Excel
            day_log_status = self._log_data_to_excel(task_start_time=self.task_start_time,
                                                          task_end_time=self.task_end_time,
                                                          work_minutes=work_minutes,
                                                          pause_minutes=pause_minutes,
                                                          total_task_minutes=total_task_minutes)

            # to update the days' work duration in the UI
            # if a new day has started, this method reads the data from the Excel freshly
            # hence calling this after logging to excel
            # we need not call this for prev_day log because - if preV-day logging is triggered by _check_for_day_change_periodically(self) i.e., when the timer is paused or running, updating UI display might confuse the user
            # so, we only call this when the same day or new day datat is logged to the Excel
            self._update_days_work_minutes_display(work_minutes)

        return day_log_status


    # if there is an error in saving the data to the Excel file (e.g., file is opened and so permission is denied), we have to stop timer and show 'Error' status
    # when user clicks stop_btn again, we have to try saving to the Excel again (e.g., user closed the file now and hit stop_btn again)
    # since we set the running status = STOPPED in the else block also to stop the timer, next time when user clicks stop_btn, we will not try to save the data
    def _end_timer(self):
        """
        Stops the timer, saves the task log to Excel, and resets the timer state
        stop timer (_update_timer())
        reset the timer text (timer_text)
        change the timer running status (is_timer_running)
        change the symbol on the start button
        deselect task in task_list_menu
        """
        # if the timer is not stopped i.e., is_timer_status is running/paused
        if self.is_timer_running != TimerStatus.STOPPED:
            if not self.auto_end:
                # _end_timer() is not triggered by the system but by the user

                # to capture when the task has ended and to be logged to the Excel
                self.task_end_time = dt.datetime.now()
                # accumulate seconds if the timer is not in paused state before ending the task
                if self.is_timer_running != TimerStatus.PAUSED:
                    self.segment_end_time = dt.datetime.now()
                    self._seconds_accumulator()

            # check for multi-day tasks and log the data to Excel
            self.is_timer_running = TimerStatus.STOPPED
            log_status = self._check_day_split_and_log()
            # log additional data for debugging
            # self._dev_log_data()

            # show status of saving the data to the Excel file
            if log_status:
                self._update_status_label("Saved", 0)
                self.auto_end = False
                self._reset_timer()
                # reset the multi-day task ID
                if self.multiday_start_date: self.multiday_start_date = None
            else:
                # if failed to save the data to the Excel
                self._update_status_label("Error", 1)
                # if we set this to STOPPED, we can't attempt to retry saving to the Excel file
                self.is_timer_running = TimerStatus.PAUSED
                self.start_btn.configure(text="▶")


    def _reset_timer(self, status=""):
        """
        Resets the timer to its initial stopped state
        stop timer
        reset seconds to 0
        change timer running status, stop button symbol
        :return:
        """
        # assume user wants to continue existing the task, but reset the timer
        # so this does not clear the selection in task_list_menu or current_task
        # if the timer is not stopped i.e., is_timer_running is running/paused

        # change the running status
        self.is_timer_running = TimerStatus.STOPPED
        self.task_list_menu.configure(state="normal")
        self.current_task=""
        # we can't edit combobox when the state is disabled
        # the state is disabled when start_btn is clicked,
        # the state is set to normal inside the reset_timer,
        # so after that we can set the value to "", else this line will have no change
        # Gemini AI or qwen did not catch this
        self.task_list_menu.set("")
        self.task_start_time = None
        self.task_end_time = None
        self.segment_start_time = None
        self.segment_end_time = None
        self.work_seconds = 0.0
        # Multi - day tracking variables reset
        self.work_seconds_logged = 0
        self.timer_text.set("00:00:00")
        self.notes_textbox.delete("1.0", "end") # clear notes
        self._show_placeholder() # show placeholder text
        self.start_btn.configure(text="▶")
        self.manage_tasks_btn.configure(command=self._manage_task_status, text_color="#4a4a4a")
        # as we are updating self.is_timer_running = TimerStatus.STOPPED, _update_timer() will stop as it runs only when the timer is running
        # self._update_timer()
        # to remove focus from notes entry field if the notes were being typed
        self.app.focus()
        # we use reset_timer method inside the end_timer method too to avoid code repetition as there are many common operations between both the methods,
        # however, the status for both the methods is diff
        # for reset_timer -> "Reset"
        # for end_timer -> "Saved"/"Error"
        # to address different status for both methods we have a check before updating status
        # we pass status "Reset" for reset_btn call
        # we pass nothing for call from end_timer
        if status:
            self._update_status_label(status, 0)
        # print("Timer reset")


    def _hide_app_window(self):
        """
        hides the app window
        used for custom_minimize button
        """
        self.app.withdraw()


    def _start_drag(self, event):
        """
        gets the initial absolute coordinates of the app window, mouse pointer wrt to the screen
        """
        # Store the initial ABSOLUTE screen coordinates of the mouse click
        self.start_mouse_x_root = event.x_root
        self.start_mouse_y_root = event.y_root

        # Store the initial ABSOLUTE screen coordinates of the app window's top-left corner
        self.start_window_x_root = self.app.winfo_x()
        self.start_window_y_root = self.app.winfo_y()


    def _do_drag(self, event):
        """
        handles repositioning of the app window on mouse hold and drag
        """
        # calculate the total displacement (change) of the mouse from its starting point
        deltax_root = event.x_root - self.start_mouse_x_root
        deltay_root = event.y_root - self.start_mouse_y_root

        # calculate the new absolute window position
        window_new_x = self.start_window_x_root + deltax_root
        window_new_y = self.start_window_y_root + deltay_root

        # reposition the app window at the new coordinates
        self.app.geometry(f"+{window_new_x}+{window_new_y}")


    def _get_resource_path(self, file_name):
        """
        Retrieve the absolute path to resource (file_name) for dev and for PyInstaller (.exe)
        This only returns the absolute file path and does not check 'if the file actually exists/valid/corrupted/readable', just path construction
        :param file_name: str file name of the asset whose path is to be retrieved
        :return: file path of the asset
        """
        if hasattr(sys, "_MEIPASS"):
            # app is running as a PyInstaller bundle if the sys._MEIPASS (Multi-Executable Installer) exists
            base_path = sys._MEIPASS
        else:
            # app is running as a script
            # base_path is the current directory where the script is running from/located in
            # asset should also be in the same directory
            base_path = os.path.abspath(".")
        # print(f"Resolved {file_name} path: {os.path.join(base_path, file_name)}")
        return os.path.join(base_path, file_name)


    #---------system tray icon [start]---------

    def _get_icon(self, icon_name) -> ImageFile:
        """
        checks the existence of icon at the path returned by _get_resource_path() method
        and if the icon file is valid, readable, not corrupted
        if it exists and valid, returns the icon
        else returns a fallback icon created with ImageDraw
        :param str icon_name
        :return: icon_image
        """
        # get the app icon path
        app_icon_path = self._get_resource_path(icon_name)
        # print(f"{app_icon_path=}")

        if os.path.exists(app_icon_path):
            icon_image = Image.open(app_icon_path)
        else:
            # if any error with app_icon, return a blank image with app initials

            # create a blank image with a blue background
            icon_image = Image.new(mode="RGB", size=(36, 36), color="#05428b")
            # create an image drawer object to write app name 'TK' for timekeeper to the blank image
            drawer = ImageDraw.Draw(icon_image)
            # get a font to draw the app initials into blank image
            # drawer.getfont() -> this gives 'self._draw(no_color_updates=True) # faster drawing without color changes'
            # image_font = ImageDraw.Draw(Image.new("RGB", (1, 1))).getfont()
            # draw app initials on to the blank image
            drawer.text((10, 10), text="TK", fill="white")

        return icon_image


    def _initialize_systray_icon(self):
        """
        Initializes a python sys tray (pystray) icon
        :return:
        """

        # 1. get the path of the sys tray icon image file -> _get_resource_path()
        # 2. check if the image file exists and is valid at the path -> _get_systray_icon()
        # 3. create menu items -> pystray.MenuItem()
        # 4. create sys tray icon -> pystray.Icon()
        # 5. run the sys tray icon in a loop on a separate thread -> self.systray_thread = threading.Thread()
        try:
            # create menu items for the sys tray icon right-click
            # default True to make it the default action on single click with LMB on the sys tray icon
            menu_items = (
                pystray.MenuItem(f"Open {self.app_title}", self._show_app_window, default=True),
                pystray.MenuItem("Hide", self._hide_app_window),
                pystray.MenuItem("Quit", self._quit_app)
            )

            # get the image file to use as icon
            icon_image = self._get_icon(self.app_icon)
            # create the systray icon with pystray
            # name="time_keeper_widget" -> used by the os/pyinstaller
            # icon_image -> icon shown in sys tray
            # f"{self.app_title} Widget" -> title/tooltip that shows when mouse is hovered on the icon
            self.systray_icon = pystray.Icon("time_keeper_widget", icon_image, f"{self.app_title} Widget", menu=menu_items)
            # start the sys tray icon loop
            self.systray_icon.run_detached()
        except Exception as e:
            print(f"FATAL ERROR: System tray icon creation/run failed: {e}")
            self.app.destroy()


    def _show_app_window(self):
        """
        Opens the app window and brings to the front
        :return:
        """

        # check if the app window is visible or not
        is_app_visible = self.app.winfo_ismapped()

        if not is_app_visible:
            self.app.deiconify()

        self.app.after(0, lambda: self.app.attributes('-topmost', True))
        self.app.after(10, lambda: self.app.attributes('-topmost', False))  # Release after 500ms (150+500)


    def _quit_app(self):
        """
        Quit the app entirely
        """

        # stop the timer
        self._end_timer()

        # cleanups to ensure no memory leaks
        if self.status_update_queue:
            self.app.after_cancel(self.status_update_queue)

        if self.timer_running_queue:
            self.app.after_cancel(self.timer_running_queue)

        if self.check_day_change_queue:
            self.app.after_cancel(self.check_day_change_queue)

        # stop the system tray icon
        if self.systray_icon:
            self.systray_icon.stop()
            self.systray_icon = None

        # destroy ctk window and exit the mainloop
        self.app.destroy()
        sys.exit(0)

    # ---------system tray icon [end]---------

    def _open_excel_file(self):
        """
        Opens the Excel file using the default system application
        Provides user feedback via the status label
        """
        if not os.path.exists(self.excel_file):
            # if the file does not exist
            self._update_status_label("Error", 1)
            return

        try:
            if sys.platform.startswith('win'):
                # Windows: uses the default application for the file type
                os.startfile(self.excel_file)
                self._update_status_label("Open", 0)
            elif sys.platform.startswith('darwin'):
                # macOS: uses the 'open' command
                import subprocess
                subprocess.run(['open', self.excel_file], check=True)
                self._update_status_label("Opened", 0)
            elif sys.platform.startswith('linux'):
                # Linux: uses 'xdg-open' which opens with the default app
                import subprocess
                subprocess.run(['xdg-open', self.excel_file], check=True)
                self._update_status_label("Opened", 0)
            else:
                self._update_status_label("Error", 1)

        except (FileNotFoundError, Exception) as e:
            # This might happen if the command itself (e.g., 'open', 'xdg-open') is not found
            self._update_status_label("Error", 1)
            print(f"Error opening the file: {e}")


    def _open_about(self):
        about_window = ctk.CTkToplevel(self.app)
        about_window.title("About")
        # about_window.geometry("260x260")
        about_window.resizable(False, False)

        # Set the toolbar icon
        icon_path = self._get_resource_path("app_icon.ico")
        if os.path.exists(icon_path):
            about_window.wm_iconbitmap(icon_path)
            about_window.after(200, lambda: about_window.iconbitmap(icon_path))
        else:
            print(f"Icon file not found for About window: {icon_path}")

        about_window.grid_columnconfigure(1, weight=1)

        app_icon_image = ctk.CTkImage(light_image=self._get_icon("app_icon.ico"), dark_image=self._get_icon("app_icon.ico"), size=(48,48))
        app_icon_label = ctk.CTkLabel(about_window, text="", image=app_icon_image)
        app_icon_label.grid(row=1, column=1, padx=30, pady=(30, 5), sticky="we")


        app_name_label = ctk.CTkLabel(about_window, text=self.app_title, text_color="#6f7a83", font=ctk.CTkFont(weight="bold", size=16))
        app_name_label.grid(row=2, column=1, padx=30, sticky="we")

        description_text = "A lightweight & minimalistic app for time tracking.\nNo accounts, no fuss - just focus!"
        description_label = ctk.CTkLabel(about_window, text=description_text, text_color="#6f7a83",
                                      font=ctk.CTkFont(size=12), wraplength=190)
        description_label.grid(row=3, column=1, padx=20, pady=(12,0), sticky="we")

        brief_steps = ctk.CTkLabel(about_window, text="Add Task -> Track Time -> Log to Excel",
                                   font=ctk.CTkFont(size=12), text_color="#6f7a83")
        brief_steps.grid(row=4, column=1, sticky="we")

        author_label = ctk.CTkLabel(about_window, text="By Akshay (@akshay_r2 on X)", text_color="#6f7a83",
                                      font=("Segoe UI", 12, "bold"))
        author_label.grid(row=5, column=1, padx=30, pady=(12, 5), sticky="we")


    def _manage_task_status(self):
        """
        Opens a new window to manage the status of the tasks
        Only 'Active' tasks are shown in the task_list_menu ComboBox via _get_task_list() method
        :return:
        """
        # 1. Create a new window to manage task status
        manage_window = ctk.CTkToplevel(self.app)
        manage_window.title("Manage Tasks")
        manage_window.resizable(False, False)

        # Set the toolbar icon
        icon_path = self._get_resource_path("app_icon.ico")
        if os.path.exists(icon_path):
            manage_window.wm_iconbitmap(icon_path)
            manage_window.after(200, lambda: manage_window.iconbitmap(icon_path))
        else:
            print(f"Icon file not found for Manage Tasks window: {icon_path}")

        # 3. Save changes
        def save_changes_to_task_status():
            # 1. Get the active tasks after user changes
            active_tasks = [cb.cget("text") for cb in task_checkboxes if cb.get() == "on"]

            # 2. Change the status in the dict as per user changes
            for task_item in self.all_tasks_dict_list:
                # Excel column name will be the key for each dict
                if task_item[self.tasks_col_name] in active_tasks:
                    task_item["Status"] = self.task_active_status_symbol
                else:
                    task_item["Status"] = ""

            # 4. Write the data back to the Excel with openpyxl as pd rewrites entire Excel
            save_error_message = ""
            try:
                wb = load_workbook(self.excel_file)
                if self.excel_tasks_sheet not in wb.sheetnames:
                    save_error_message= f"Error: {self.excel_tasks_sheet} sheet not found in '{self.excel_file}'. No update performed."
                    return

                ws = wb[self.excel_tasks_sheet]
                # Remove existing task rows excluding the header row
                ws.delete_rows(2, ws.max_row - 1)

                # Prepare the rows and write to the Excel
                # Order of data in rows - Task	Status	Added_On
                for task_dict in self.all_tasks_dict_list:
                    row_values = [
                        task_dict.get(self.tasks_col_name, ""),
                        task_dict.get("Status", ""),
                        task_dict.get("Added_On", "")
                    ]
                    ws.append(row_values)

                wb.save(self.excel_file)

                # 5. Refresh the task list for ComboBox dropdown and close the manage task status window if save is successful

                # update the task list dropdown ComboBox
                self.task_list = self._get_task_list()
                # to ensure "<Add new task...>" is at the top of the list
                self.task_list[1:] = sorted(self.task_list[1:])
                # update the combobox with the new task_list
                self.task_list_menu.configure(values=self.task_list)

                # Reset the current task if any as it may be made inactive only if the timer is not running
                # If the timer is running or paused, and if we reset the current_task, task name in Excel log will be blank
                if self.is_timer_running == TimerStatus.STOPPED:
                    self.current_task = ""
                    self.task_list_menu.set("")
                manage_window.destroy()

            except (pd.errors.ParserError, FileNotFoundError, InvalidFileException, Exception) as err:
                print(err)
                save_error_message = "Error on save. Please try again :("
                save_error_label.configure(text=save_error_message)


        # 2. Show the tasks with checkbox if the tasks exist, else show the error message
        if self.all_tasks_dict_list:
            scrollable_frame = ctk.CTkScrollableFrame(manage_window, width=220, height=350)
            scrollable_frame.grid(row=1, column=1, padx=10, pady=(10,5))

            # create checkbox for each task
            task_checkboxes = [] # holds the created ctk CheckBoxes

            # sort the tasks in alphabetical order of task name
            sorted_tasks_dict_list = sorted(self.all_tasks_dict_list,
                                            key=lambda task_item: task_item[self.tasks_col_name])

            for item in sorted_tasks_dict_list:
                checked_state = "on" if item["Status"] == self.task_active_status_symbol else "off"
                checked_state_var = ctk.StringVar(value=checked_state)
                task_checkbox = ctk.CTkCheckBox(scrollable_frame, text=item[self.tasks_col_name],
                                                variable=checked_state_var, onvalue="on", offvalue="of",
                                                corner_radius=4, border_width=2, fg_color="#085bbe",
                                                hover_color="#05428b", font=("Segoe UI", 14))
                # wrap lengthy task names
                task_checkbox._text_label.configure(wraplength=280)
                task_checkboxes.append(task_checkbox)
                task_checkbox.grid(row=len(task_checkboxes), column=1, padx=10, pady=8, sticky="we")

            # 3. Save the changes
            # to show error warning if any on saving the changes
            save_error_label = ctk.CTkLabel(manage_window, text="️", text_color="#afb4ba")
            save_error_label.grid(row=2, column=1, sticky="we")
            save_btn = ctk.CTkButton(manage_window, text="Save", fg_color="#085bbe", hover_color="#05428b",
                                     command=save_changes_to_task_status)
            save_btn.grid(row=3, column=1, pady=(0,12))
        else:
            # if no tasks were found or error on file read
            manage_window.geometry("250x420")
            error_label = ctk.CTkLabel(manage_window, text_color="#4f575d", text="No Tasks Found :(", font=("Segoe UI", 16, "bold"))
            # weights to center the label inside the window
            manage_window.grid_columnconfigure(1, weight=1)
            manage_window.grid_rowconfigure(1, weight=1)
            error_label.grid(row=1, column=1, sticky="nswe")


    def _build_ui(self):
        """
        To build the widgets of the app
        """
        # toolbar to show app name, days work duration, and minimize button
        toolbar_frame = ctk.CTkFrame(self.app, height=30, fg_color="#2c2c2c", corner_radius=0)
        toolbar_frame.grid(row=1, column=1, columnspan=3, sticky="we")

        title_label = ctk.CTkLabel(toolbar_frame, text=self.app_title, font=ctk.CTkFont(size=12))
        # adjusted 'pady' to ensure title, days work duration are aligned horizontally
        title_label.grid(row=1, column=1, sticky="w", pady=(7,3), padx=10)

        # get the duration in hh:mm format for display - text_color= #575f66
        days_work_minutes_formated = self._humanize_time(self.days_work_minutes)
        self.days_work_label = ctk.CTkLabel(toolbar_frame, text=f"Day: {days_work_minutes_formated}",
                                       text_color="#4f575d", font=("Segoe UI", 13, "bold"))
        self.days_work_label.grid(row=1, column=2, sticky="e")

        custom_minimize_btn = ctk.CTkButton(toolbar_frame, text="\u2013", fg_color="#343638",
                                            hover_color="#585a5c", width=40, height=20,
                                            font=("Segoe UI Symbol", 15),
                                            command=self._hide_app_window)
        custom_minimize_btn.grid(row=1, column=3,  padx=(10,5), pady=5)
        # to ensure the custom_minimize button sticks to the right edge
        toolbar_frame.grid_columnconfigure(2, weight=1)

        # binding the mouse events to enable dragging functionality to the toolbar_frame and title_label
        toolbar_frame.bind("<Button-1>", self._start_drag)
        toolbar_frame.bind("<B1-Motion>", self._do_drag)
        
        title_label.bind("<Button-1>", self._start_drag)
        title_label.bind("<B1-Motion>", self._do_drag)

        self.days_work_label.bind("<Button-1>", self._start_drag)
        self.days_work_label.bind("<B1-Motion>", self._do_drag)

        # dropdown menu to choose the tasks from task_list
        self.task_list_menu = ctk.CTkComboBox(self.app, values=self.task_list, command=self._list_menu_callback)
        self.task_list_menu.grid(row=2, column=1, padx=10, pady=(10, 0), sticky="ew", columnspan=3)
        # remove the default option displayed from combobox dropdown (defaults to the first option)
        self.task_list_menu.set("")
        # to add a new task on press of the Enter key
        self.task_list_menu.bind("<Return>", self._add_task_on_enter)

        # hint text to show how to add a new task to the task_list
        hint_label = ctk.CTkLabel(self.app, text="Type new task & press Enter", font=("Segoe UI", 12, "bold"), height=5, text_color="#7a848d")
        hint_label.grid(row=3, column=1, columnspan=3, padx=10, sticky="w")

        # status text to show a message on task addition
        self.status_label = ctk.CTkLabel(self.app, text="", font=("Segoe UI", 12, "bold"), height=5)
        self.status_label.grid(row=3, column=2, columnspan=3, sticky="e", padx=(0, 11))

        # entry widget to display the running timer
        # set initial text
        initial_timer_text= "00:00:00"
        self.timer_text.set(initial_timer_text)
        self.timer_display = ctk.CTkEntry(self.app, textvariable=self.timer_text, height=70,
                                     font=("Segoe UI Symbol", 40, "bold"), justify="center", state="disabled",
                                     text_color="#9e9e9e")
        self.timer_display.grid(row=4, column=1, columnspan=3, padx=10, pady=10, sticky="we")

        # field to enter notes for the task (removed width = 220)
        self.notes_textbox = ctk.CTkTextbox(self.app, text_color="#f2f2f2", height=62,
                                          font=("Segoe UI", 14), wrap="word",
                                          border_width=1, border_color="#4c5154")
        self.notes_textbox.grid(row=5, column=1, columnspan=3, sticky="we", padx=(10,8))

        # show placeholder text at the start
        self._show_placeholder()
        # bind focus-in and focus-out events to handle show/hide of the placeholder text
        self.notes_textbox.bind("<FocusIn>", self._notes_focus_in)
        self.notes_textbox.bind("<FocusOut>", self._notes_focus_out)

        # button frame to hold the buttons and adjust their spacing and widths
        # we have a separate frame for buttons as we have to place 4 buttons in 3 columns
        buttons_frame = ctk.CTkFrame(self.app, fg_color="transparent", height=30)
        buttons_frame.grid(row=6, column=1, columnspan=3, sticky="we", pady=(10,0))

        # column weight to ensure the reset button takes more space
        buttons_frame.grid_columnconfigure(3, weight=2)

        # buttons to control the functionality
        # self.start_btn is an instance variable as the text changes ▶ -> ⏸ in run_timer method
        self.start_btn = ctk.CTkButton(buttons_frame, text="▶", command=self._run_timer, width=49,
                                       font=("Segoe UI Symbol", 16, "bold"), fg_color="#085bbe",
                                       hover_color="#05428b")
        self.start_btn.grid(padx=(10,8), row=1, column=1, sticky="we")

        end_btn = ctk.CTkButton(buttons_frame, text="⏹", width=49, fg_color="#085bbe",
                                font=("Segoe UI Symbol", 16, "bold"), hover_color="#05428b",
                                command=self._end_timer)
        end_btn.grid(row=1, column=2, sticky="we")

        reset_btn = ctk.CTkButton(buttons_frame, text="Reset", width=60, fg_color="#242424", hover_color="#414449",
                                  border_color="#414449", border_width=1,
                                  command=lambda: self._reset_timer("Reset"))
        reset_btn.grid(padx=(8,5), row=1, column=3, sticky="we")

        excel_btn_icon = ctk.CTkImage(light_image=self._get_icon(self.excel_btn_icon),
                                      dark_image=self._get_icon(self.excel_btn_icon), size=(19,19))
        open_excel_btn = ctk.CTkButton(buttons_frame, image=excel_btn_icon, fg_color="#242424",
                                       border_color="#414449", border_width=0, hover_color="#414449", width=1,
                                       text="", command=self._open_excel_file)
        open_excel_btn.grid(row=1, column=4, padx=(0,10), sticky="w")

        signature_label = ctk.CTkLabel(self.app, text="akshay;)", text_color="#262626",  height=5,
                                         font=ctk.CTkFont(size=8, weight="bold", slant="italic"))
        signature_label.grid(row=7, column=3, sticky="se")
        # 575f66 4f575d
        about_btn = ctk.CTkButton(self.app, text="About", height=1, fg_color="transparent", width=1,
                                         hover_color="#2c2c2c", text_color="#4a4a4a", font=ctk.CTkFont(weight="bold"), command=self._open_about)
        about_btn.grid(row=7, column=1, padx=(8,6), pady=3, sticky="we")

        # Made as an instance variable to  control enabled state
        self.manage_tasks_btn = ctk.CTkButton(self.app, text="Manage Tasks", height=1,
                                              fg_color="transparent", width=1, hover_color="#2c2c2c",
                                              text_color="#4a4a4a", font=ctk.CTkFont( weight="bold"),
                                              command=self._manage_task_status)
        self.manage_tasks_btn.grid(row=7, column=2, pady=3, sticky="we")

        quit_btn = ctk.CTkButton(self.app, text="Quit", height=1, fg_color="transparent", width=1,
                                         hover_color="#2c2c2c", text_color="#4a4a4a", font=ctk.CTkFont(weight="bold"), command=self._quit_app)
        quit_btn.grid(row=7, column=3, padx=(8,10), pady=3, sticky="we")


    def _get_dpi_scaling(self, hwnd):
        """
        Returns the DPI scaling factor for the given window handle.
        Example: 1.0 for 100%, 1.25 for 125%, 1.5 for 150%, etc.
        """
        try:
            dpi = windll.user32.GetDpiForWindow(hwnd)
            return dpi / 96.0
        except Exception as e:
            print(f"Error getting DPI: {e}")
            return 1.0


    def position_window(self):
        # -----------positioning window in the bottom right corner of the screen-----------
        self.app.update_idletasks()

        # Logical vs. Physical Pixels: Screen operates at a physical pixel resolution of 1920x1080, but with 150% DPI scaling, Windows presents a "logical" resolution 1280x720 to applications
        # The coordinates passed to geometry() for positioning might be interpreted as physical pixels by the OS, even if Tkinter is internally working with logical pixels.
        # Solution: By taking the logical positions, multiplying them by dpi_scale_factor (1.5), and then passing those x_pos_physical and y_pos_physical values to app.geometry(), we ensure the window is placed at the correct physical coordinates - bottom right corner

        # position the app window in the bottom right corner of the screen with a margin
        right_margin = 5
        # separate bottom margin to offset the taskbar height
        bottom_margin = 85
        # by trial and error, found app wxh = 240x280 as app.winfo_width() and app.winfo_height() were not giving correct dimensions, probably due to DPI scaling
        app_width = 240
        app_height = 280

        # Get window handle for DPI detection
        hwnd = self.app.winfo_id()
        # 1.5
        dpi_scale_factor = self._get_dpi_scaling(hwnd)
        # print(f"{dpi_scale_factor=}")

        screen_width_logical = self.app.winfo_screenwidth()
        screen_height_logical = self.app.winfo_screenheight()

        # calculate the logical x y coordinates
        x_logical = screen_width_logical - app_width - right_margin
        y_logical = screen_height_logical - app_height - bottom_margin

        # convert logical coordinates to physical for .geometry()
        x_physical = int(x_logical * dpi_scale_factor)
        y_physical = int(y_logical * dpi_scale_factor)

        # set the app window's position
        self.app.geometry(f"+{x_physical}+{y_physical}")


app = TaskTimer()
